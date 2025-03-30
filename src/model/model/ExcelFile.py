import os
import zipfile
from typing import List
import pandas as pd
import tempfile

from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from pathlib import Path


class ExcelFile:
    DEFAULT_COLUMN_INDEX: int = 3  # Константа для столбца по умолчанию

    def __init__(self, file_path: str) -> None:
        """
        Инициализация класса для обработки Excel файла.

        Args:
            file_path (str): Путь к файлу Excel.

        Raises:
            FileNotFoundError: Если файл не существует.
            ValueError: Если файл не является валидным .xlsx или повреждён.
        """
        self.file_path: Path = Path(file_path)
        self._validate_file()

    def _validate_file(self) -> None:
        """
        Проверяет, является ли файл валидным .xlsx.

        Raises:
            FileNotFoundError: Если файл не существует.
            ValueError: Если файл не .xlsx, повреждён или не может быть открыт.
        """
        if not self.file_path.exists():
            raise FileNotFoundError(f"Файл {self.file_path} не найден.")
        if self.file_path.suffix.lower() != ".xlsx":
            raise ValueError(f"Файл {self.file_path} не является файлом .xlsx.")

        # Проверяем возможность открыть файл через openpyxl
        try:
            load_workbook(self.file_path)
        except zipfile.BadZipFile:
            raise ValueError(f"Файл {self.file_path} повреждён или не является архивом.")
        except Exception as e:
            if 'sharedStrings.xml' in str(e):
                self.file_path = Path(self._fix_shared_strings_case())
            else:
                raise ValueError(f"Не удалось открыть файл {self.file_path}: {e}")

    def _fix_shared_strings_case(self) -> str:
        """
        Исправляет регистр файла SharedStrings.xml в архиве Excel файла.

        Returns:
            str: Путь к исправленному файлу.

        Raises:
            FileNotFoundError: Если SharedStrings.xml не найден в архиве.
            IOError: Если не удалось создать исправленный файл.
        """
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path: Path = Path(temp_dir)
            fixed_file_path: Path = self.file_path.with_name(f"fixed_{self.file_path.name}")

            try:
                with zipfile.ZipFile(self.file_path, 'r') as zip_ref:
                    zip_ref.extractall(temp_path)
            except zipfile.BadZipFile:
                raise ValueError(f"Не удалось извлечь содержимое архива {self.file_path}.")

            shared_strings_path: Path | None = None
            for root, _, files in os.walk(temp_path):
                for file in files:
                    if file.lower() == 'sharedstrings.xml':
                        shared_strings_path = Path(root) / file
                        shared_strings_path.rename(shared_strings_path.parent / 'sharedStrings.xml')
                        break

            if not shared_strings_path:
                raise FileNotFoundError("Файл SharedStrings.xml не найден в архиве.")

            try:
                with zipfile.ZipFile(fixed_file_path, 'w') as zip_ref:
                    for foldername, _, filenames in os.walk(temp_path):
                        for filename in filenames:
                            file_path = os.path.join(foldername, filename)
                            arcname = os.path.relpath(file_path, temp_path)
                            zip_ref.write(file_path, arcname)
            except IOError as e:
                raise IOError(f"Ошибка при создании исправленного файла {fixed_file_path}: {e}")

        return str(fixed_file_path)

    def _extract_images(self) -> List[Image]:
        """
        Извлекает изображения из Excel файла.

        Returns:
            List[Image]: Список объектов изображений.

        Raises:
            ValueError: Если в файле нет изображений.
            RuntimeError: Если не удалось загрузить файл.
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active  # Выбираем активный лист
        except Exception as e:
            raise RuntimeError(f"Ошибка при загрузке файла {self.file_path}: {e}")

        # Извлекаем все изображения
        images: List[Image] = [item for item in ws._images if isinstance(item, Image)]
        if not images:
            raise ValueError("В файле нет изображений.")
        return images

    def _extract_column_text(self, column_index: int = DEFAULT_COLUMN_INDEX) -> List[str]:
        """
        Извлекает текстовые данные из указанного столбца Excel файла.

        Args:
            column_index (int, optional): Индекс столбца для извлечения текста. По умолчанию 3.

        Returns:
            List[str]: Список строк из указанного столбца.

        Raises:
            IndexError: Если указанный столбец выходит за границы таблицы.
            ValueError: Если не удалось прочитать файл.
        """
        try:
            data: pd.DataFrame = pd.read_excel(self.file_path)
        except Exception as e:
            raise ValueError(f"Не удалось прочитать Excel файл {self.file_path}: {e}")

        if column_index >= data.shape[1]:
            raise IndexError(
                f"Указанный столбец с индексом {column_index} выходит за границы таблицы. "
                f"Всего столбцов: {data.shape[1]}"
            )
        column_data = data.iloc[:, column_index].tolist()
        return [str(i) for i in column_data if isinstance(i, str) or pd.notna(i)]

    @staticmethod
    def _save_images_with_names(images: List[Image], text_list: List[str], output_folder: str) -> None:
        """
        Сохраняет изображения с именами, соответствующими строкам текста.

        Args:
            images (List[Image]): Список изображений.
            text_list (List[str]): Список строк текста.
            output_folder (str): Путь к папке для сохранения изображений.

        Raises:
            ValueError: Если количество изображений и строк текста не совпадает.
            RuntimeError: Если не удалось извлечь данные изображения.
            OSError: Если не удалось создать папку или сохранить файл.
        """
        if len(images) != len(text_list):
            raise ValueError(f"Количество изображений ({len(images)}) и строк текста ({len(text_list)}) не совпадает!")

        try:
            os.makedirs(output_folder, exist_ok=True)
        except OSError as e:
            raise OSError(f"Не удалось создать папку {output_folder}: {e}")

        for image, text in zip(images, text_list):
            # Создаем безопасное имя файла
            safe_text: str = "".join(c if c.isalnum() else "_" for c in text)
            image_filename: str = os.path.join(output_folder, f'{safe_text}.png')

            # Сохраняем изображение
            try:
                with open(image_filename, 'wb') as img_file:
                    img_file.write(image._data())  # Извлечение бинарного содержимого
            except AttributeError:
                raise RuntimeError(f"Не удалось извлечь данные изображения для {image_filename}.")
            except OSError as e:
                raise OSError(f"Ошибка при сохранении изображения {image_filename}: {e}")

    def process_images_and_text(self, output_folder: str, column_index: int = DEFAULT_COLUMN_INDEX) -> None:
        """
        Извлекает изображения и текст, проверяет соответствие, сохраняет изображения с именами строк текста.

        Args:
            output_folder (str): Путь к папке для сохранения изображений.
            column_index (int, optional): Индекс столбца для извлечения текста. По умолчанию 3.

        Raises:
            RuntimeError: Если произошла ошибка при обработке файла.
        """
        try:
            images: List[Image] = self._extract_images()
            text_list: List[str] = self._extract_column_text(column_index)
            self._save_images_with_names(images, text_list, output_folder)
        except Exception as e:
            raise RuntimeError(f"Ошибка при обработке файла {self.file_path}: {e}")